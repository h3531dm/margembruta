"""
Conciliação de Vendas — TCPOS x Opera
"""

import re
import io
from collections import defaultdict

import pandas as pd
import pdfplumber
import streamlit as st

st.set_page_config(
    page_title="Conciliação de Vendas",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+3:wght@400;500;600;700&family=IBM+Plex+Mono:wght@500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'Source Sans 3', sans-serif;
    font-size: 15px;
}

.stApp { background: #e8ecf0 !important; }

.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 2rem !important;
    max-width: 100% !important;
}
[data-testid="stAppViewBlockContainer"] { padding-top: 0 !important; }

/* ── FAIXA SUPERIOR: branca com título azul, sem espaço extra ── */
header[data-testid="stHeader"] {
    background: #ffffff !important;
    border-bottom: 3px solid #1c3557 !important;
    height: auto !important;
    min-height: 0 !important;
}
header[data-testid="stHeader"]::before {
    content: "CONCILIAÇÃO DE VENDAS";
    display: block;
    padding: 0.85rem 2.5rem;
    font-family: 'Source Sans 3', sans-serif;
    font-size: 1.05rem;
    font-weight: 700;
    color: #1c3557;
    letter-spacing: 0.4px;
    text-transform: uppercase;
    line-height: 1;
}
/* Esconde botões padrão do header */
header[data-testid="stHeader"] > div:last-child { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }

/* ── CARD BRANCO ── */
.gov-card {
    background: #ffffff;
    border-radius: 10px;
    border: 1px solid #d4dae2;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    padding: 1.75rem 2rem;
    margin-bottom: 1rem;
}
.gov-card-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: #1c3557;
    margin-bottom: 1.25rem;
    padding-bottom: 0.75rem;
    border-bottom: 1px solid #e8ecf0;
}

/* ── BOTÃO AZUL DECORATIVO ── */
.upload-btn {
    width: 100%;
    padding: 0.9rem 1rem;
    background: #1c3557;
    color: #ffffff !important;
    border-radius: 8px;
    font-size: 0.92rem;
    font-weight: 600;
    text-align: center;
    cursor: pointer;
    margin-bottom: 6px;
    pointer-events: none; /* visual apenas; o clique vai para o uploader abaixo */
}

/* ── UPLOADER NATIVO: INVISÍVEL (branco sobre branco) ── */
[data-testid="stFileUploader"] label { display: none !important; }

[data-testid="stFileUploadDropzone"] {
    background: #ffffff !important;
    border: none !important;
    box-shadow: none !important;
    padding: 2px 0 !important;
    min-height: 0 !important;
}

/* Ícone da nuvem — esconde */
[data-testid="stFileUploadDropzone"] svg {
    display: none !important;
}

/* Texto "Drag and drop file here" — substitui por português */
[data-testid="stFileUploadDropzone"] span {
    font-size: 0 !important;
    color: transparent !important;
}
[data-testid="stFileUploadDropzone"] span::after {
    content: "Arraste o arquivo aqui";
    font-size: 0.9rem !important;
    color: #6b7a8d !important;
}

/* Texto "Limit 200MB" — esconde */
[data-testid="stFileUploadDropzone"] small {
    display: none !important;
}

/* Botão "Browse files" — substitui por "Anexar arquivo" */
[data-testid="stFileUploadDropzone"] button {
    background: #ffffff !important;
    color: transparent !important;
    border: 1px solid #d4dae2 !important;
    box-shadow: none !important;
    font-size: 0 !important;
    padding: 0.35rem 0.9rem !important;
    border-radius: 6px !important;
    cursor: pointer !important;
    position: relative;
}
[data-testid="stFileUploadDropzone"] button::after {
    content: "Anexar arquivo";
    font-size: 0.82rem;
    font-weight: 500;
    color: #1c3557;
    position: absolute;
    left: 50%;
    top: 50%;
    transform: translate(-50%, -50%);
    white-space: nowrap;
}

/* Mantém a área clicável mas invisível */
[data-testid="stFileUploaderDropzoneInput"] {
    opacity: 0.01 !important;
    cursor: pointer !important;
}

/* ── CARD ARQUIVO CARREGADO ── */
.upload-ok {
    border: 1.5px solid #86efac;
    background: #f0fdf4;
    border-radius: 8px;
    padding: 0.8rem 1rem;
    display: flex;
    align-items: center;
    gap: 0.75rem;
    margin-top: 4px;
}
.uok-check { font-size: 1rem; color: #16a34a; }
.uok-name  { font-size: 0.85rem; font-weight: 600; color: #166534; }
.uok-tag   { font-size: 0.71rem; color: #16a34a; }

/* ── MÉTRICAS ── */
.metrics-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.75rem;
}
.metric-box {
    border: 1px solid #d4dae2;
    border-radius: 8px;
    padding: 0.9rem 1rem;
    background: #fff;
}
.metric-val {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.75rem;
    font-weight: 700;
    line-height: 1;
    margin-bottom: 0.25rem;
}
.metric-lbl {
    font-size: 0.72rem;
    font-weight: 600;
    color: #7a8799;
    text-transform: uppercase;
    letter-spacing: 0.4px;
}
.m-ok  .metric-val { color: #1a7f3c; }
.m-spl .metric-val { color: #b45309; }
.m-div .metric-val { color: #b91c1c; }
.m-tc  .metric-val { color: #c2410c; }
.m-op  .metric-val { color: #6d28d9; }

/* ── TOTAIS ── */
.totais-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.75rem;
}
.total-box {
    border: 1px solid #d4dae2;
    border-radius: 8px;
    padding: 0.85rem 1rem;
    background: #f8fafc;
}
.total-val {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.05rem;
    font-weight: 600;
    color: #1c3557;
    margin-bottom: 0.2rem;
}
.total-lbl {
    font-size: 0.68rem;
    font-weight: 600;
    color: #7a8799;
    text-transform: uppercase;
    letter-spacing: 0.4px;
}

/* ── DIFERENÇA ── */
.diff-box {
    border-radius: 8px;
    padding: 0.85rem 1.1rem;
    margin-top: 0.75rem;
}
.diff-ok   { background: #f0fdf4; border: 1px solid #86efac; }
.diff-warn { background: #fef2f2; border: 1px solid #fca5a5; }
.diff-lbl  { font-size: 0.68rem; font-weight: 600; color: #7a8799;
             text-transform: uppercase; letter-spacing: 0.4px; margin-bottom: 0.3rem; }
.diff-val  { font-family: 'IBM Plex Mono', monospace; font-size: 1.15rem; font-weight: 700; }
.diff-pos  { color: #15803d; }
.diff-neg  { color: #b91c1c; }
.diff-zero { color: #4b5563; }

/* ── ALERTAS ── */
.gov-alert {
    border-radius: 8px;
    padding: 0.7rem 1rem;
    font-size: 0.84rem;
    font-weight: 500;
    margin-bottom: 1rem;
}
.alert-ok  { background: #f0fdf4; color: #166534; border: 1px solid #86efac; }
.alert-err { background: #fef2f2; color: #991b1b; border: 1px solid #fca5a5; }

/* ── BOTÃO EXPORT ── */
.stDownloadButton > button {
    background: #1c3557 !important;
    color: #fff !important;
    border: none !important;
    border-radius: 6px !important;
    font-family: 'Source Sans 3', sans-serif !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    padding: 0.5rem 1.4rem !important;
}
.stDownloadButton > button:hover { background: #26527e !important; }

/* ── TABELA ── */
.stDataFrame { border-radius: 8px !important; border: 1px solid #d4dae2 !important; }

/* ── MULTISELECT ── */
[data-baseweb="select"] > div {
    border-color: #d4dae2 !important;
    border-radius: 7px !important;
    background: #fff !important;
}

/* ── TABS DEBUG ── */
.stTabs [data-baseweb="tab-list"] {
    background: #edf0f4 !important;
    border-radius: 7px !important;
    padding: 3px !important;
}
.stTabs [data-baseweb="tab"] {
    font-size: 0.8rem !important;
    color: #6b7a8d !important;
    border-radius: 5px !important;
}
.stTabs [aria-selected="true"] {
    background: #fff !important;
    color: #1c3557 !important;
    font-weight: 700 !important;
}

section[data-testid="stSidebar"] {
    background: #fff !important;
    border-right: 1px solid #d4dae2;
}
.stNumberInput input {
    border: 1px solid #d4dae2 !important;
    border-radius: 6px !important;
}
.stCaption { color: #9098a8 !important; font-size: 0.74rem !important; }

/* ── BOX TRANSACTION CODES — compacto ── */
.tc-inline {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    margin-top: 0.75rem;
    padding-top: 0.75rem;
    border-top: 1px solid #e8ecf0;
}
.tc-label {
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.6px;
    text-transform: uppercase;
    color: #9098a8;
    white-space: nowrap;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# EXTRAÇÃO OPERA
# ─────────────────────────────────────────────────────────

def extrair_opera(file):
    """
    Extrai vendas do Opera por NF da linha CHECK#.

    Retorna:
      vendas    {nf: valor_liquido}   — já descontados estornos
      contagem  {nf: qtd_linhas}
      desc      {nf: descrição}       — ex: "2060 Bar 1 Alimento Cafe"
      cancelados set(nf)              — NFs cujo valor líquido é zero (cancelados)
    """
    # Acumula valores brutos (positivos e negativos separados)
    bruto_pos  = defaultdict(float)   # soma dos lançamentos positivos
    bruto_neg  = defaultdict(float)   # soma dos lançamentos negativos (em abs)
    contagem   = defaultdict(int)
    desc       = {}                   # primeira descrição encontrada por NF
    linhas     = []

    with pdfplumber.open(file) as pdf:
        for p in pdf.pages:
            txt = p.extract_text()
            if txt:
                for l in txt.splitlines():
                    l = l.strip()
                    if l:
                        linhas.append(l)

    extrair_opera._linhas = linhas

    re_nf    = re.compile(r'NF:(\d+)', re.IGNORECASE)
    re_valor = re.compile(r'(-?\d+\.\d{2})')       # captura negativos também
    re_desc  = re.compile(r'\d{4}\s+.+')            # padrão "2060 Bar 1 ..."

    for i, linha in enumerate(linhas):
        if 'CHECK#' not in linha:
            continue
        nfs = re_nf.findall(linha)
        if not nfs or i == 0:
            continue
        nf       = nfs[-1]
        anterior = linhas[i - 1]

        # Extrai valor (pode ser negativo = estorno)
        vals = re_valor.findall(anterior)
        if not vals:
            continue

        # Pega o primeiro valor não-zero (positivo ou negativo)
        valor = None
        for v in vals:
            f = float(v)
            if abs(f) > 0:
                valor = f
                break
        if valor is None:
            continue

        if valor > 0:
            bruto_pos[nf] += valor
        else:
            bruto_neg[nf] += abs(valor)

        contagem[nf] += 1

        # Captura descrição da linha anterior (primeira vez que aparece este NF)
        if nf not in desc:
            # Procura padrão "DDDD Texto Texto" na linha anterior
            m = re_desc.search(anterior)
            if m:
                desc[nf] = m.group(0).strip()

    # Calcula valor líquido: positivo - negativo
    vendas     = {}
    cancelados = set()
    for nf in set(bruto_pos) | set(bruto_neg):
        liquido = round(bruto_pos.get(nf, 0) - bruto_neg.get(nf, 0), 2)
        vendas[nf] = liquido
        if abs(liquido) < 0.01:
            cancelados.add(nf)

    # Extrai data do relatório — busca "From Date DD/MM/YY"
    data_relatorio = ""
    re_data = re.compile(r'From Date\s+(\d{2}/\d{2}/\d{2,4})', re.IGNORECASE)
    for linha in linhas:
        m = re_data.search(linha)
        if m:
            raw = m.group(1)  # ex: 17/04/26
            partes = raw.split("/")
            if len(partes) == 3:
                d, mes, a = partes
                ano = f"20{a}" if len(a) == 2 else a
                data_relatorio = f"{d}/{mes}/{ano}"
            break

    extrair_opera._dict          = vendas
    extrair_opera._cancelados    = cancelados
    extrair_opera._desc          = desc
    extrair_opera._data          = data_relatorio
    return vendas, dict(contagem), desc, cancelados


# ─────────────────────────────────────────────────────────
# EXTRAÇÃO TCPOS
# ─────────────────────────────────────────────────────────

def extrair_tcpos(file):
    """
    Extrai vendas do TCPOS com tratamento correto de cancelamentos.

    Formatos de valor encontrados no PDF real:
      $1,532.35      → venda normal   → positivo
      ($1,532.35)    → cancelamento   → negativo (subtrair)

    Lógica:
      - Filtra linhas que começam com HH:MM
      - NF = coluna índice 2
      - Detecta se o valor está entre parênteses → negativo
      - Calcula valor LÍQUIDO por NF (soma - estornos)
      - NFs com líquido zero vão para cancelados_tcpos
    """
    bruto_pos = defaultdict(float)
    bruto_neg = defaultdict(float)
    linhas    = []

    with pdfplumber.open(file) as pdf:
        for p in pdf.pages:
            txt = p.extract_text()
            if txt:
                for l in txt.splitlines():
                    l = l.strip()
                    if l:
                        linhas.append(l)

    extrair_tcpos._linhas = linhas

    re_hora         = re.compile(r'^\d{2}:\d{2}\s')
    # Captura valor positivo: $1,532.35
    re_valor_pos    = re.compile(r'(?<!\()\$([0-9,]+\.\d{2})(?!\))')
    # Captura valor negativo (entre parênteses): ($1,532.35)
    re_valor_neg    = re.compile(r'\(\$([0-9,]+\.\d{2})\)')

    for linha in linhas:
        # Ignora linhas de cabeçalho, totais e rodapé
        if any(k in linha for k in ['Total', '#', 'Time', 'NF', 'Check', 'Impresso', 'Printed', 'UNIDADE', 'Hora', 'Lojas', 'Data:', 'CUPONS']):
            continue
        if not re_hora.match(linha):
            continue
        partes = linha.split()
        if len(partes) < 5:
            continue

        nf = partes[2]

        # Verifica primeiro se é cancelamento (parênteses)
        m_neg = re_valor_neg.search(linha)
        if m_neg:
            try:
                valor = float(m_neg.group(1).replace(',', ''))
                bruto_neg[nf] += valor
                continue
            except ValueError:
                pass

        # Valor positivo normal
        m_pos = re_valor_pos.search(linha)
        if m_pos:
            try:
                valor = float(m_pos.group(1).replace(',', ''))
                bruto_pos[nf] += valor
            except ValueError:
                pass

    # Calcula líquido por NF
    vendas            = {}
    cancelados_tcpos  = set()
    for nf in set(bruto_pos) | set(bruto_neg):
        liquido = round(bruto_pos.get(nf, 0) - bruto_neg.get(nf, 0), 2)
        vendas[nf] = liquido
        if abs(liquido) < 0.01:
            cancelados_tcpos.add(nf)

    extrair_tcpos._dict            = vendas
    extrair_tcpos._cancelados      = cancelados_tcpos
    extrair_tcpos._bruto_pos       = dict(bruto_pos)
    extrair_tcpos._bruto_neg       = dict(bruto_neg)
    return vendas


# ─────────────────────────────────────────────────────────
# CONCILIAÇÃO
# ─────────────────────────────────────────────────────────

def conciliar(tcpos, opera_vals, opera_cnt, opera_desc,
              opera_cancelados, tcpos_cancelados, tol=0.01):
    """
    Conciliação bidirecional com detecção de cancelamentos em ambos os sistemas.

    NF cancelado = valor líquido zero (positivos - estornos = 0).
    Opera e TCPOS tratados de forma simétrica.
    """
    rows = []
    todos_nfs = sorted(set(tcpos) | set(opera_vals))

    for nf in todos_nfs:
        vt  = tcpos.get(nf)
        vo  = opera_vals.get(nf)
        qt  = opera_cnt.get(nf, 0)
        des = opera_desc.get(nf, "")

        # ── Cancelado em ambos ──
        if nf in opera_cancelados and nf in tcpos_cancelados:
            rows.append({
                "NF": nf, "Descrição Opera": des,
                "Valor TCPOS": 0.0, "Valor Opera": 0.0,
                "Diferença": 0.0, "Linhas Opera": qt,
                "Status": "Cancelado nos dois",
            })
            continue

        # ── Cancelado só no Opera ──
        if nf in opera_cancelados:
            rows.append({
                "NF": nf, "Descrição Opera": des,
                "Valor TCPOS": vt, "Valor Opera": 0.0,
                "Diferença": vt if vt else 0.0,
                "Linhas Opera": qt, "Status": "Cancelado no Opera",
            })
            continue

        # ── Cancelado só no TCPOS ──
        if nf in tcpos_cancelados:
            rows.append({
                "NF": nf, "Descrição Opera": des,
                "Valor TCPOS": 0.0, "Valor Opera": vo,
                "Diferença": -(vo or 0), "Linhas Opera": qt,
                "Status": "Cancelado no TCPOS",
            })
            continue

        # ── Conciliação normal ──
        if vt is not None and vo is not None:
            dif    = round(vt - vo, 2)
            status = ("Split" if qt > 1 else "OK") if abs(dif) <= tol else "Divergente"
        elif vt is not None:
            dif, status = vt,  "Só TCPOS"
        else:
            dif, status = -vo, "Só Opera"

        rows.append({
            "NF": nf, "Descrição Opera": des,
            "Valor TCPOS": vt, "Valor Opera": vo,
            "Diferença": dif, "Linhas Opera": qt, "Status": status,
        })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────

def gerar_excel(df, data_relatorio=""):
    """
    Gera Excel com:
    - Data do relatório no topo (linha 1)
    - Cores por status
    - AutoFilter nativo (filtro já ativo ao abrir)
    - Coluna Descrição Opera
    - Coluna Status é a última (facilita filtrar)
    """
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Escreve os dados a partir da linha 3 (deixa 2 linhas para o cabeçalho)
        df.to_excel(writer, index=False, sheet_name="Conciliação", startrow=2)
        ws = writer.sheets["Conciliação"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        num_cols = len(df.columns)
        last_col = get_column_letter(num_cols)

        # ── LINHA 1: data do relatório ──────────────────────
        ws.merge_cells(f"A1:{last_col}1")
        cel_data = ws["A1"]
        label = f"Conciliação de Vendas — {data_relatorio}" if data_relatorio else "Conciliação de Vendas"
        cel_data.value     = label
        cel_data.font      = Font(bold=True, color="FFFFFF", size=12)
        cel_data.fill      = PatternFill("solid", fgColor="1C3557")
        cel_data.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── LINHA 2: vazia (espaçamento) ────────────────────
        ws.row_dimensions[2].height = 6

        cores = {
            "OK":                  "D1FAE5",
            "Split":               "FEF3C7",
            "Divergente":          "FEE2E2",
            "Só TCPOS":            "FFEDD5",
            "Só Opera":            "EDE9FE",
            "Cancelado no Opera":  "F1F5F9",
            "Cancelado no TCPOS":  "F1F5F9",
            "Cancelado nos dois":  "F1F5F9",
        }

        thin = Side(style="thin", color="D1D5DB")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── LINHA 3: cabeçalho das colunas ─────────────────
        for cell in ws[3]:
            cell.font      = Font(bold=True, color="1C3557", size=10)
            cell.fill      = PatternFill("solid", fgColor="DBEAFE")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = brd

        # Índice da coluna Status (última coluna, 0-based)
        status_col_idx = len(df.columns) - 1

        # ── LINHAS DE DADOS: a partir da linha 4 ────────────
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            status_val = str(row[status_col_idx].value or "")
            cor  = cores.get(status_val, "FFFFFF")
            fill = PatternFill("solid", fgColor=cor)
            for cell in row:
                cell.fill      = fill
                cell.border    = brd
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Largura das colunas
        for ci, col in enumerate(ws.columns, 1):
            ml = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 50)

        # Altura das linhas de dados
        for row in ws.iter_rows(min_row=4):
            ws.row_dimensions[row[0].row].height = 20

        # ── AUTOFILTER na linha de cabeçalho (linha 3) ──────
        ws.auto_filter.ref = f"A3:{last_col}{ws.max_row}"

        # Congela até a linha 3 (data + cabeçalho sempre visíveis)
        ws.freeze_panes = "A4"

    return out.getvalue()


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────

def brl(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

def dif_fmt(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    s = f"R$ {abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    if v > 0.01:  return f"+ {s} no TCPOS"
    if v < -0.01: return f"- {s} no Opera"
    return "R$ 0,00"

def render_diff(tt, to):
    d = round(tt - to, 2)
    if abs(d) < 0.01:
        return '<div class="diff-box diff-ok"><div class="diff-lbl">Diferença total entre sistemas</div><div class="diff-val diff-zero">Sem diferença</div></div>'
    s = f"R$ {abs(d):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    if d > 0:
        return f'<div class="diff-box diff-warn"><div class="diff-lbl">Diferença total entre sistemas</div><div class="diff-val diff-pos">+ {s} no TCPOS</div></div>'
    return f'<div class="diff-box diff-warn"><div class="diff-lbl">Diferença total entre sistemas</div><div class="diff-val diff-neg">- {s} no Opera</div></div>'


# ─────────────────────────────────────────────────────────
# INTERFACE
# ─────────────────────────────────────────────────────────

def main():

    with st.sidebar:
        st.markdown("**Configurações**")
        tolerancia = st.number_input("Tolerância (R$)", 0.0, 1.0, 0.01, 0.01)
        debug      = st.checkbox("Modo debug")

    _, col_mid, _ = st.columns([1, 8, 1])

    with col_mid:

        # ── CARD UPLOAD ───────────────────────────────────
        st.markdown('<div class="gov-card">', unsafe_allow_html=True)
        st.markdown('<div class="gov-card-title">Selecione os relatórios do dia</div>',
                    unsafe_allow_html=True)

        col1, col2 = st.columns(2, gap="large")

        with col1:
            # Botão azul decorativo (visual)
            st.markdown('<div class="upload-btn">Relatório Fin01106 — Opera</div>',
                        unsafe_allow_html=True)
            # Uploader real — invisível (branco sobre branco do card)
            arq_opera = st.file_uploader(
                "opera", type="pdf", key="opera", label_visibility="collapsed"
            )
            if arq_opera:
                kb = round(arq_opera.size / 1024, 1)
                st.markdown(f"""
                <div class="upload-ok">
                    <span class="uok-check">✓</span>
                    <div>
                        <div class="uok-name">{arq_opera.name}</div>
                        <div class="uok-tag">{kb} KB · carregado</div>
                    </div>
                </div>""", unsafe_allow_html=True)

        with col2:
            st.markdown('<div class="upload-btn">Relatório de Cupons Fiscais — TCPOS</div>',
                        unsafe_allow_html=True)
            arq_tcpos = st.file_uploader(
                "tcpos", type="pdf", key="tcpos", label_visibility="collapsed"
            )
            if arq_tcpos:
                kb = round(arq_tcpos.size / 1024, 1)
                st.markdown(f"""
                <div class="upload-ok">
                    <span class="uok-check">✓</span>
                    <div>
                        <div class="uok-name">{arq_tcpos.name}</div>
                        <div class="uok-tag">{kb} KB · carregado</div>
                    </div>
                </div>""", unsafe_allow_html=True)


        # ── TRANSACTION CODES — compacto dentro do card ──
        import streamlit.components.v1 as components
        components.html("""
        <style>
          .tc-row {
            display: flex;
            align-items: center;
            gap: 0.6rem;
            margin-top: 0.6rem;
            padding-top: 0.6rem;
            border-top: 1px solid #e8ecf0;
            font-family: 'Source Sans 3', sans-serif;
          }
          .tc-lbl {
            font-size: 0.67rem;
            font-weight: 700;
            letter-spacing: 0.6px;
            text-transform: uppercase;
            color: #9098a8;
            white-space: nowrap;
          }
          .tc-btn {
            padding: 0.2rem 0.65rem;
            background: #fff;
            border: 1px solid #d4dae2;
            border-radius: 5px;
            font-size: 0.72rem;
            font-weight: 600;
            color: #1c3557;
            cursor: pointer;
            transition: background 0.15s;
            white-space: nowrap;
          }
          .tc-btn:hover { background: #eef3f9; }
        </style>
        <div class="tc-row">
          <span class="tc-lbl">Transaction Codes — FIN01106</span>
          <button class="tc-btn" onclick="
            navigator.clipboard.writeText('2303,2302,2423,2301,2422,2421,2420,2320,2441,2440,2316,2436,2323,2322,2443,2321,2442,2342,2341,2340,2336,2456,2346,2345,2343,2001,2000,2356,2012,3101,2011,3100,2010,2009,2006,2004,2003,2002,2019,2017,3106,2014,3103,3102,2040,2055,3023,3022,2052,2051,2050,2065,3393,2061,3392,2060,2057,2070,2076,2075,2071,2081,2080,2085,2079,5152,2702,2701,2700,2703,2716,2603,2602,2723,2601,2722,2600,2721,2720,2607,2611,2616,2736,2623,2622,2743,2621,2742,2620,2741,2740,2636,2756,2403,2402,2401,2643,2400,2642,2641,2640,2416,2656').then(function(){
              this.textContent = '✓ Copiado';
              var btn = this;
              setTimeout(function(){ btn.textContent = 'Copiar'; }, 1500);
            }.bind(this));
          ">Copiar</button>
        </div>
        """, height=52)

        st.markdown('</div>', unsafe_allow_html=True)

        if not arq_opera or not arq_tcpos:
            st.markdown(
                '<p style="font-size:0.83rem;color:#9098a8;margin-top:0.25rem;">'
                'Clique nos botões acima para selecionar os relatórios do dia.'
                '</p>', unsafe_allow_html=True)
            return

        # ── PROCESSAR ─────────────────────────────────────
        with st.spinner("Processando os relatórios..."):
            try:
                opera_vals, opera_cnt, opera_desc, opera_cancel = extrair_opera(arq_opera)
                tcpos         = extrair_tcpos(arq_tcpos)
                cancel_tcpos  = getattr(extrair_tcpos, "_cancelados", set())
                df            = conciliar(tcpos, opera_vals, opera_cnt, opera_desc,
                                          opera_cancel, cancel_tcpos, tolerancia)
            except Exception as e:
                st.error(f"Erro ao processar: {e}")
                return

        if df.empty:
            st.warning("Nenhum registro encontrado. Verifique os arquivos.")
            return

        sc     = df["Status"].value_counts()
        n_ok   = sc.get("OK", 0)
        n_spl  = sc.get("Split", 0)
        n_div  = sc.get("Divergente", 0)
        n_tc   = sc.get("Só TCPOS", 0)
        n_op   = sc.get("Só Opera", 0)
        n_can  = sc.get("Cancelado no Opera", 0) + sc.get("Cancelado no TCPOS", 0) + sc.get("Cancelado nos dois", 0)
        n_prob = n_div + n_tc + n_op
        tt     = df["Valor TCPOS"].sum(skipna=True)
        to     = df["Valor Opera"].sum(skipna=True)

        # ── ALERTA ────────────────────────────────────────
        cls = "alert-ok" if n_prob == 0 else "alert-err"
        msg = (f"Conciliação concluída — {len(df)} registros analisados, nenhuma pendência encontrada."
               if n_prob == 0 else
               f"{n_prob} registro(s) com pendência — revise os itens na tabela abaixo.")
        st.markdown(f'<div class="gov-alert {cls}">{msg}</div>',
                    unsafe_allow_html=True)

        # ── CARD RESUMO ───────────────────────────────────
        st.markdown('<div class="gov-card">', unsafe_allow_html=True)
        st.markdown('<div class="gov-card-title">Resumo da Conciliação</div>',
                    unsafe_allow_html=True)

        st.markdown(f"""
        <div class="metrics-grid" style="grid-template-columns: repeat(6, 1fr);">
            <div class="metric-box m-ok">
                <div class="metric-val">{n_ok}</div>
                <div class="metric-lbl">OK</div>
            </div>
            <div class="metric-box m-spl">
                <div class="metric-val">{n_spl}</div>
                <div class="metric-lbl">Split no Opera</div>
            </div>
            <div class="metric-box m-div">
                <div class="metric-val">{n_div}</div>
                <div class="metric-lbl">Valor Divergente</div>
            </div>
            <div class="metric-box m-tc">
                <div class="metric-val">{n_tc}</div>
                <div class="metric-lbl">Só no TCPOS</div>
            </div>
            <div class="metric-box m-op">
                <div class="metric-val">{n_op}</div>
                <div class="metric-lbl">Só no Opera</div>
            </div>
            <div class="metric-box" style="border-top: 3px solid #64748b;">
                <div class="metric-val" style="color:#475569;">{n_can}</div>
                <div class="metric-lbl">Cancelados</div>
            </div>
        </div>
        <div class="totais-grid" style="margin-top:0.75rem;">
            <div class="total-box">
                <div class="total-val">{len(tcpos)}</div>
                <div class="total-lbl">Registros TCPOS</div>
            </div>
            <div class="total-box">
                <div class="total-val">{len(opera_vals)}</div>
                <div class="total-lbl">Registros Opera</div>
            </div>
            <div class="total-box">
                <div class="total-val">{brl(tt)}</div>
                <div class="total-lbl">Total TCPOS</div>
            </div>
            <div class="total-box">
                <div class="total-val">{brl(to)}</div>
                <div class="total-lbl">Total Opera</div>
            </div>
        </div>
        {render_diff(tt, to)}
        """, unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        # ── CARD TABELA ───────────────────────────────────
        st.markdown('<div class="gov-card">', unsafe_allow_html=True)
        st.markdown('<div class="gov-card-title">Resultado por NF</div>',
                    unsafe_allow_html=True)

        opts   = df["Status"].unique().tolist()
        filtro = st.multiselect("Filtrar por status:", opts, default=opts)
        df_f   = df[df["Status"].isin(filtro)].copy() if filtro else df.copy()

        df_ex                = df_f.copy()
        df_ex["Valor TCPOS"] = df_ex["Valor TCPOS"].apply(brl)
        df_ex["Valor Opera"] = df_ex["Valor Opera"].apply(brl)
        df_ex["Diferença"]   = df_ex["Diferença"].apply(dif_fmt)

        st.dataframe(
            df_ex, use_container_width=True, hide_index=True,
            column_config={
                "NF":               st.column_config.TextColumn("NF",              width="small"),
                "Descrição Opera":  st.column_config.TextColumn("Descrição Opera", width="large"),
                "Valor TCPOS":      st.column_config.TextColumn("Valor TCPOS",     width="medium"),
                "Valor Opera":      st.column_config.TextColumn("Valor Opera",     width="medium"),
                "Diferença":        st.column_config.TextColumn("Diferença",       width="medium"),
                "Linhas Opera":     st.column_config.NumberColumn("Linhas Opera",  width="small"),
                "Status":           st.column_config.TextColumn("Status",          width="medium"),
            },
        )
        st.caption(f"{len(df_f)} de {len(df)} registros")
        st.markdown('</div>', unsafe_allow_html=True)

        # ── EXPORTAR ──────────────────────────────────────
        data_rel = getattr(extrair_opera, "_data", "")
        st.download_button(
            "Exportar Excel",
            data=gerar_excel(df, data_relatorio=data_rel),
            file_name="conciliacao_tcpos_opera.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # ── DEBUG ─────────────────────────────────────────
        if debug:
            st.markdown('<div class="gov-card" style="margin-top:1rem;">',
                        unsafe_allow_html=True)
            st.markdown('<div class="gov-card-title">Debug — Inspeção dos PDFs</div>',
                        unsafe_allow_html=True)
            t1, t2 = st.tabs(["Opera", "TCPOS"])
            with t1:
                ls = getattr(extrair_opera, "_linhas", [])
                st.caption(f"{len(ls)} linhas lidas")
                st.code("\n".join(ls[:40]), language="text")
                st.json(getattr(extrair_opera, "_dict", {}))
            with t2:
                ls = getattr(extrair_tcpos, "_linhas", [])
                st.caption(f"{len(ls)} linhas lidas")
                st.code("\n".join(ls[:40]), language="text")
                st.json(getattr(extrair_tcpos, "_dict", {}))
            st.markdown('</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
